import openpyxl

# adding a new row to postcode excel file
# sheet contains name, postcode
def add_to_sheet(name, postcode):
    # filling Balance Excel sheet
    xls_file = "Locations.xlsx"
    wb = openpyxl.load_workbook(xls_file)
    sheet = wb.active
    last_index = sheet.max_row + 1
    sheet.cell(last_index, 1, name)
    sheet.cell(last_index, 2, postcode)
    wb.save(xls_file)
    # update sheet
    print('Sheet')


def location(name):
    # find a specific player's current balance
    xls_file = "Locations.xls"
    wb = openpyxl.load_workbook(xls_file)
    sheet = wb.active
    postcode = -1
    # using iteration in the location table to find the name's location
    for i in range(2, sheet.max_row + 1):
        if str(sheet.cell(i, 1).value) == name:
            postcode = int(sheet.cell(i, 5).value)
    return postcode

